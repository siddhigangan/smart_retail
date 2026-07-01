import os
import csv
import io
import time
from datetime import datetime
from decimal import Decimal
from sqlalchemy import func
from sqlalchemy.orm import Session
from app.models.product import Product
from app.models.shelf import Shelf
from app.models.inventory_mapping import InventoryMapping
from app.models.adjustment_log import AdjustmentLog
from app.models.refill_log import RefillLog
from app.models.bill import BillItem

import openpyxl

class ProductManagementService:
    @staticmethod
    def get_shelf_recommendations(db: Session, category: str) -> list[dict]:
        """
        Queries shelves matching category and returns recommendations sorted by remaining capacity.
        """
        # Fetch all shelves
        shelves = db.query(Shelf).all()
        
        # Calculate sum of current shelf quantities per shelf
        shelf_sums = db.query(
            InventoryMapping.shelf_id,
            func.sum(InventoryMapping.current_shelf_quantity)
        ).group_by(InventoryMapping.shelf_id).all()
        
        shelf_qty_map = {row[0]: int(row[1]) for row in shelf_sums}
        
        results = []
        for s in shelves:
            # Match category case-insensitively
            is_match = (s.category.strip().lower() == category.strip().lower()) if category else False
            
            current_qty = shelf_qty_map.get(s.id, 0)
            # Default physical capacity of any shelf is 100 units
            remaining_cap = max(0, 100 - current_qty)
            
            status = "🟢 Recommended" if is_match and remaining_cap > 0 else ("🟡 Full" if remaining_cap <= 0 else "⚪ Available")
            
            results.append({
                "id": s.id,
                "floor_number": s.floor_number,
                "aisle": s.aisle,
                "rack": s.rack,
                "shelf_number": s.shelf_number,
                "category": s.category,
                "remaining_capacity": remaining_cap,
                "status": status,
                "score": (2 if is_match else 0) + (1 if remaining_cap > 0 else -10) # scoring for sorting
            })
            
        # Sort by match score first, then remaining capacity descending
        results = sorted(results, key=lambda x: (x["score"], x["remaining_capacity"]), reverse=True)
        return results

    @staticmethod
    def get_download_template() -> str:
        """
        Generates CSV template headers for bulk import.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Barcode", "Product Name", "Brand", "Category", "Sub Category",
            "Supplier", "MRP", "Selling Price", "Cost Price", "GST", "HSN",
            "Quantity", "Reorder Level", "Shelf Capacity", "Warehouse Quantity"
        ])
        # Sample row
        writer.writerow([
            "8901234567890", "Sample Chocolate Bar", "ChocoMart", "Biscuits", "Chocolates",
            "Nagpur Traders", "50.00", "45.00", "35.00", "18.00", "1806",
            "120", "20", "20", "100"
        ])
        return output.getvalue()

    @staticmethod
    def validate_product_data(
        db: Session,
        barcode: str,
        name: str,
        category: str,
        price: float,
        cost_price: float,
        quantity: int,
        reorder_level: int,
        shelf_id: int = None,
        existing_barcodes_in_import: set = None
    ) -> list[str]:
        """
        Performs thorough validation checks and returns list of error messages.
        """
        errors = []
        if not barcode or not barcode.strip():
            errors.append("Barcode is required.")
        else:
            # Check db duplicate
            db_dup = db.query(Product).filter(Product.barcode == barcode.strip()).first()
            if db_dup:
                errors.append(f"Barcode '{barcode}' already exists in database.")
            # Check in-file duplicate
            if existing_barcodes_in_import is not None and barcode.strip() in existing_barcodes_in_import:
                errors.append(f"Barcode '{barcode}' is duplicated in the uploaded file.")

        if not name or not name.strip():
            errors.append("Product Name is required.")
            
        if not category or not category.strip():
            errors.append("Category is required.")

        try:
            p_val = float(price)
            if p_val < 0:
                errors.append("Selling Price cannot be negative.")
        except (ValueError, TypeError):
            errors.append("Invalid Selling Price numeric value.")

        if cost_price is not None:
            try:
                cp_val = float(cost_price)
                if cp_val < 0:
                    errors.append("Cost Price cannot be negative.")
            except (ValueError, TypeError):
                errors.append("Invalid Cost Price numeric value.")

        try:
            q_val = int(quantity)
            if q_val < 0:
                errors.append("Total Quantity cannot be negative.")
        except (ValueError, TypeError):
            errors.append("Invalid Total Quantity integer value.")

        try:
            r_val = int(reorder_level)
            if r_val < 0:
                errors.append("Reorder Level cannot be negative.")
        except (ValueError, TypeError):
            errors.append("Invalid Reorder Level integer value.")

        return errors

    @staticmethod
    def manual_add_product(db: Session, data: dict) -> Product:
        """
        Creates a product, binds inventory mapping, and links selected shelf.
        """
        barcode = data["barcode"].strip()
        name = data["name"].strip()
        category = data["category"].strip()
        selling_price = Decimal(str(data["selling_price"]))
        cost_price = Decimal(str(data["cost_price"])) if data.get("cost_price") else None
        mrp = Decimal(str(data["mrp"])) if data.get("mrp") else None
        gst = Decimal(str(data["gst"])) if data.get("gst") else None
        quantity = int(data["quantity"])
        reorder_level = int(data.get("reorder_level", 0))

        # Check duplicate
        db_dup = db.query(Product).filter(Product.barcode == barcode).first()
        if db_dup:
            raise ValueError(f"Product with barcode '{barcode}' already exists.")

        # Create Product
        prod = Product(
            barcode=barcode,
            name=name,
            brand=data.get("brand"),
            category=category,
            sub_category=data.get("sub_category"),
            description=data.get("description"),
            unit=data.get("unit"),
            pack_size=data.get("pack_size"),
            supplier=data.get("supplier"),
            mrp=mrp,
            price=selling_price,
            cost_price=cost_price,
            gst=gst,
            hsn_code=data.get("hsn_code"),
            quantity=quantity,
            reorder_level=reorder_level,
            expiry_required=data.get("expiry_required", False),
            expiry_date=datetime.strptime(data["expiry_date"], "%Y-%m-%d").date() if data.get("expiry_date") else None
        )
        db.add(prod)
        db.flush()

        # Shelf mapping parameters
        shelf_id = int(data["shelf_id"])
        shelf_capacity = int(data.get("shelf_capacity", 20))
        current_shelf_qty = int(data.get("current_shelf_quantity", min(quantity, shelf_capacity)))
        warehouse_qty = int(data.get("warehouse_quantity", quantity - current_shelf_qty))
        minimum_shelf_qty = int(data.get("minimum_shelf_quantity", 5))

        # Create Mapping
        mapping = InventoryMapping(
            product_id=prod.id,
            shelf_id=shelf_id,
            shelf_capacity=shelf_capacity,
            current_shelf_quantity=current_shelf_qty,
            warehouse_quantity=warehouse_qty,
            minimum_shelf_quantity=minimum_shelf_qty
        )
        db.add(mapping)
        db.commit()
        return prod

    @staticmethod
    def parse_import_file(file_content: bytes, filename: str) -> list[dict]:
        """
        Parses CSV or Excel uploaded file and returns rows list.
        """
        rows = []
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = wb.active
            # Get headers from first row
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            for row_idx in range(2, sheet.max_row + 1):
                row_vals = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
                if not any(row_vals):
                    continue # skip blank row
                
                # Make dict matching CSV headers
                row_dict = {}
                for idx, header in enumerate(headers):
                    val = row_vals[idx]
                    row_dict[header] = str(val).strip() if val is not None else ""
                rows.append(row_dict)
        else:
            # CSV decoding
            content_str = file_content.decode("utf-8-sig", errors="ignore")
            reader = csv.DictReader(io.StringIO(content_str))
            for row in reader:
                if not any(row.values()):
                    continue
                rows.append(row)
                
        return rows

    @classmethod
    def validate_and_preview_import(cls, db: Session, rows: list[dict]) -> dict:
        """
        Validates the rows list and separates them into Valid, Duplicate, and Invalid.
        """
        valid_records = []
        duplicate_records = []
        invalid_records = []
        
        seen_barcodes = set()

        for idx, row in enumerate(rows):
            barcode = row.get("Barcode", "").strip()
            name = row.get("Product Name", "").strip()
            brand = row.get("Brand", "").strip()
            category = row.get("Category", "").strip()
            sub_category = row.get("Sub Category", "").strip()
            supplier = row.get("Supplier", "").strip()
            mrp = row.get("MRP", "").strip()
            selling_price = row.get("Selling Price", "").strip()
            cost_price = row.get("Cost Price", "").strip()
            gst = row.get("GST", "").strip()
            hsn = row.get("HSN", "").strip()
            quantity = row.get("Quantity", "").strip()
            reorder_level = row.get("Reorder Level", "").strip()
            shelf_cap = row.get("Shelf Capacity", "").strip()
            wh_qty = row.get("Warehouse Quantity", "").strip()

            # Format default variables for validation checks
            price_val = 0.0
            cost_val = None
            qty_val = 0
            reorder_val = 0

            try: price_val = float(selling_price) if selling_price else 0.0
            except: pass
            try: cost_val = float(cost_price) if cost_price else None
            except: pass
            try: qty_val = int(quantity) if quantity else 0
            except: pass
            try: reorder_val = int(reorder_level) if reorder_level else 0
            except: pass

            errors = cls.validate_product_data(
                db=db,
                barcode=barcode,
                name=name,
                category=category,
                price=price_val,
                cost_price=cost_val,
                quantity=qty_val,
                reorder_level=reorder_val,
                existing_barcodes_in_import=seen_barcodes
            )

            record_info = {
                "row_index": idx + 2,
                "barcode": barcode,
                "name": name,
                "category": category,
                "quantity": qty_val,
                "price": price_val,
                "row_data": row,
                "errors": errors
            }

            if errors:
                # Differentiate duplicate in database/file vs invalid
                is_dup = any("already exists" in e or "duplicated in the uploaded file" in e for e in errors)
                if is_dup:
                    duplicate_records.append(record_info)
                else:
                    invalid_records.append(record_info)
            else:
                valid_records.append(record_info)
                seen_barcodes.add(barcode)

        return {
            "valid": valid_records,
            "duplicates": duplicate_records,
            "invalid": invalid_records
        }

    @classmethod
    def execute_import(cls, db: Session, valid_records: list[dict]) -> dict:
        """
        Performs the database writes of valid records.
        Suggests shelves and performs quantity splits if not explicitly provided.
        """
        start_time = time.time()
        imported = 0
        skipped = 0
        errors = 0

        for rec in valid_records:
            row = rec["row_data"]
            try:
                barcode = row.get("Barcode", "").strip()
                name = row.get("Product Name", "").strip()
                category = row.get("Category", "").strip()
                
                selling_price = Decimal(row.get("Selling Price", "0.00").strip() or "0.00")
                cost_price = Decimal(row.get("Cost Price", "0.00").strip()) if row.get("Cost Price") else None
                mrp = Decimal(row.get("MRP", "0.00").strip()) if row.get("MRP") else None
                gst = Decimal(row.get("GST", "0.00").strip()) if row.get("GST") else None
                quantity = int(row.get("Quantity", "0").strip() or "0")
                reorder_level = int(row.get("Reorder Level", "0").strip() or "0")

                # Suggest best shelf for Category
                recoms = cls.get_shelf_recommendations(db, category)
                if not recoms:
                    # Fallback: get any shelf in db
                    fallback_shelf = db.query(Shelf).first()
                    if not fallback_shelf:
                        raise ValueError("No shelves configured in system. Please add shelves first.")
                    shelf_id = fallback_shelf.id
                else:
                    shelf_id = recoms[0]["id"] # highest remaining recommendation score

                # Get shelf capacity and split quantities
                shelf_capacity = int(row.get("Shelf Capacity", "20").strip() or "20")
                
                # Split rules
                if quantity > shelf_capacity:
                    current_shelf_qty = shelf_capacity
                    warehouse_qty = quantity - shelf_capacity
                else:
                    current_shelf_qty = quantity
                    warehouse_qty = 0

                # Override if provided in import file
                if row.get("Warehouse Quantity"):
                    try:
                        warehouse_qty = int(row.get("Warehouse Quantity").strip())
                        current_shelf_qty = quantity - warehouse_qty
                    except:
                        pass

                # Create Product
                prod = Product(
                    barcode=barcode,
                    name=name,
                    brand=row.get("Brand"),
                    category=category,
                    sub_category=row.get("Sub Category"),
                    supplier=row.get("Supplier"),
                    mrp=mrp,
                    price=selling_price,
                    cost_price=cost_price,
                    gst=gst,
                    hsn_code=row.get("HSN"),
                    quantity=quantity,
                    reorder_level=reorder_level
                )
                db.add(prod)
                db.flush()

                # Create Inventory Mapping
                mapping = InventoryMapping(
                    product_id=prod.id,
                    shelf_id=shelf_id,
                    shelf_capacity=shelf_capacity,
                    current_shelf_quantity=current_shelf_qty,
                    warehouse_quantity=warehouse_qty,
                    minimum_shelf_quantity=5
                )
                db.add(mapping)
                imported += 1
            except Exception as e:
                errors += 1
                db.rollback()
                continue
        
        db.commit()
        duration = round(time.time() - start_time, 2)
        
        return {
            "imported": imported,
            "skipped": skipped,
            "duplicates": 0,
            "errors": errors,
            "duration_seconds": duration
        }

    @staticmethod
    def bulk_stock_update(db: Session, updates_list: list[dict]) -> dict:
        """
        Accepts list of barcode, current_quantity, new_quantity, reason.
        Updates DB quantity, adjusts warehouse stock parts, and creates AdjustmentLog audits.
        """
        updated = 0
        failed = 0

        for entry in updates_list:
            try:
                barcode = entry["barcode"].strip()
                new_qty = int(entry["new_quantity"])
                reason = entry.get("reason", "Bulk Stock Update").strip()

                if new_qty < 0:
                    failed += 1
                    continue

                prod = db.query(Product).filter(Product.barcode == barcode).first()
                if not prod:
                    failed += 1
                    continue

                before_qty = prod.quantity
                delta = new_qty - before_qty

                # Update product total quantity
                prod.quantity = new_qty

                # Adjust inventory mapping warehouse quantity accordingly
                mapping = db.query(InventoryMapping).filter(InventoryMapping.product_id == prod.id).first()
                if mapping:
                    # Apply delta to warehouse stock
                    mapping.warehouse_quantity = max(0, mapping.warehouse_quantity + delta)

                # Create Adjustment Log
                log = AdjustmentLog(
                    product_id=prod.id,
                    barcode=barcode,
                    quantity_changed=delta,
                    before_quantity=before_qty,
                    after_quantity=new_qty,
                    reason=reason
                )
                db.add(log)
                updated += 1
            except Exception as e:
                failed += 1
                continue

        db.commit()
        return {
            "updated": updated,
            "failed": failed
        }

    @staticmethod
    def adjust_inventory(db: Session, barcode: str, quantity_changed: int, reason: str) -> dict:
        """
        Adjusts product stock by quantity_changed (positive/negative) and logs it.
        """
        prod = db.query(Product).filter(Product.barcode == barcode).first()
        if not prod:
            raise ValueError(f"Product with barcode '{barcode}' not found.")

        before_qty = prod.quantity
        after_qty = max(0, before_qty + quantity_changed)
        actual_delta = after_qty - before_qty

        # Update product quantity
        prod.quantity = after_qty

        # Adjust warehouse quantity in mapping
        mapping = db.query(InventoryMapping).filter(InventoryMapping.product_id == prod.id).first()
        if mapping:
            mapping.warehouse_quantity = max(0, mapping.warehouse_quantity + actual_delta)

        # Log adjustment
        log = AdjustmentLog(
            product_id=prod.id,
            barcode=barcode,
            quantity_changed=actual_delta,
            before_quantity=before_qty,
            after_quantity=after_qty,
            reason=reason
        )
        db.add(log)
        db.commit()

        return {
            "barcode": barcode,
            "before_quantity": before_qty,
            "after_quantity": after_qty,
            "quantity_changed": actual_delta,
            "reason": reason
        }

    @staticmethod
    def get_product_profile(db: Session, product_id: int) -> dict:
        """
        Fetches the complete product details along with linked mappings, histories, and logs.
        """
        prod = db.query(Product).filter(Product.id == product_id).first()
        if not prod:
            return None

        # Fetch mapping and shelf
        mapping = db.query(InventoryMapping).filter(InventoryMapping.product_id == prod.id).first()
        shelf_info = None
        if mapping:
            shelf = db.query(Shelf).filter(Shelf.id == mapping.shelf_id).first()
            if shelf:
                shelf_info = {
                    "id": shelf.id,
                    "floor_number": shelf.floor_number,
                    "aisle": shelf.aisle,
                    "rack": shelf.rack,
                    "shelf_number": shelf.shelf_number,
                    "category": shelf.category,
                    "shelf_capacity": mapping.shelf_capacity,
                    "current_shelf_quantity": mapping.current_shelf_quantity,
                    "warehouse_quantity": mapping.warehouse_quantity,
                    "minimum_shelf_quantity": mapping.minimum_shelf_quantity
                }

        # Refill history
        refills = db.query(RefillLog).filter(RefillLog.product_id == prod.id).order_by(RefillLog.refilled_at.desc()).limit(15).all()
        refill_history = [{
            "quantity_moved": r.quantity_moved,
            "before_quantity": r.before_quantity,
            "after_quantity": r.after_quantity,
            "refilled_at": r.refilled_at.strftime("%Y-%m-%d %I:%M %p")
        } for r in refills]

        # Adjustments history
        adjusts = db.query(AdjustmentLog).filter(AdjustmentLog.product_id == prod.id).order_by(AdjustmentLog.adjusted_at.desc()).limit(15).all()
        adjustment_history = [{
            "quantity_changed": a.quantity_changed,
            "before_quantity": a.before_quantity,
            "after_quantity": a.after_quantity,
            "reason": a.reason,
            "adjusted_at": a.adjusted_at.strftime("%Y-%m-%d %I:%M %p")
        } for a in adjusts]

        # Sales history (purchases inside checkout bills)
        sales = db.query(BillItem).filter(BillItem.product_id == prod.id).order_by(BillItem.id.desc()).limit(15).all()
        sales_history = [{
            "bill_id": s.bill_id,
            "quantity": s.quantity,
            "unit_price": float(s.unit_price),
            "subtotal": float(s.subtotal)
        } for s in sales]

        return {
            "id": prod.id,
            "barcode": prod.barcode,
            "name": prod.name,
            "brand": prod.brand or "N/A",
            "category": prod.category,
            "sub_category": prod.sub_category or "N/A",
            "description": prod.description or "N/A",
            "unit": prod.unit or "N/A",
            "pack_size": prod.pack_size or "N/A",
            "supplier": prod.supplier or "N/A",
            "price": float(prod.price),
            "cost_price": float(prod.cost_price) if prod.cost_price else 0.0,
            "mrp": float(prod.mrp) if prod.mrp else 0.0,
            "gst": float(prod.gst) if prod.gst else 0.0,
            "hsn_code": prod.hsn_code or "N/A",
            "quantity": prod.quantity,
            "reorder_level": prod.reorder_level,
            "expiry_date": prod.expiry_date.strftime("%Y-%m-%d") if prod.expiry_date else "N/A",
            "shelf": shelf_info,
            "refill_history": refill_history,
            "adjustment_history": adjustment_history,
            "sales_history": sales_history
        }
